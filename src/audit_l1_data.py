"""Read-only inventory for the L1 battery-data attachments.

Creates a compact JSON evidence file from the problem statement, XLSX exports,
and MATLAB v7.3 (HDF5) container metadata.  It deliberately never writes to
the source attachments.
"""
from __future__ import annotations

import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

try:
    import h5py
except ModuleNotFoundError:  # MATLAB performs the v7.3 metadata inventory.
    h5py = None
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "data_audit"
NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_safe(value):
    if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
        return None
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [json_safe(item) for item in value]
    if isinstance(value, dict):
        return {key: json_safe(item) for key, item in value.items()}
    return value


def read_docx_text(path: Path) -> str:
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
    lines = []
    for paragraph in root.findall(".//w:p", NS):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", NS))
        if text.strip():
            lines.append(text)
    return "\n".join(lines)


def xlsx_inventory(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        sample = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
        populated_rows = sum(1 for row in ws.iter_rows(values_only=True) if any(value is not None for value in row))
        headers = [str(value) if value is not None else None for value in (sample[0] if sample else [])]
        sheets.append({
            "name": ws.title,
            "rows": populated_rows,
            "columns": ws.max_column or max((len(row) for row in sample), default=0),
            "header": headers,
            "first_rows": sample[1:4],
        })
    wb.close()
    return {"path": path.name, "bytes": path.stat().st_size, "sha256": sha256(path), "sheets": sheets}


def mat_inventory(path: Path) -> dict:
    if h5py is None:
        matlab_metadata_path = OUT / "mat_metadata.json"
        if matlab_metadata_path.exists():
            metadata = json.loads(matlab_metadata_path.read_text(encoding="utf-8"))
            source = next(item for item in metadata["sources"] if item["mat_file"] == path.name)
            return {
                "path": path.name,
                "bytes": path.stat().st_size,
                "sha256": sha256(path),
                "status": "audited_by_matlab_metadata_script",
                "batch_date": source["batch_date"],
                "batch_records": source["batch_records"],
                "top_level_variables": source["top_level_variables"],
                "record_fields": source["record_fields"],
                "cycle_fields": source["cycle_fields"],
                "summary_fields": source["summary_fields"],
                "vdlin_length": source["vdlin_length"],
            }
        return {"path": path.name, "bytes": path.stat().st_size, "sha256": sha256(path), "status": "pending_matlab_metadata"}
    datasets = []
    with h5py.File(path, "r") as handle:
        def visitor(name: str, obj: object) -> None:
            if isinstance(obj, h5py.Dataset):
                datasets.append({
                    "path": name,
                    "shape": list(obj.shape),
                    "dtype": str(obj.dtype),
                    "bytes": int(obj.size * obj.dtype.itemsize),
                    "is_reference_array": bool(h5py.check_dtype(ref=obj.dtype)),
                })
        handle.visititems(visitor)
        top_level = sorted(handle.keys())
    return {
        "path": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
        "hdf5_top_level": top_level,
        "dataset_count": len(datasets),
        "datasets": datasets,
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    docx = ROOT / "B.docx"
    result = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "scope": "read-only audit inventory; no source attachment was modified",
        "problem_statement": {"path": docx.name, "bytes": docx.stat().st_size, "sha256": sha256(docx), "text": read_docx_text(docx)},
        "xlsx_exports": [xlsx_inventory(ROOT / f"data_{index}.xlsx") for index in range(1, 4)],
        "mat_sources": [mat_inventory(ROOT / f"data_{index}.mat") for index in range(1, 4)],
    }
    result = json_safe(result)
    (OUT / "inventory.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str, allow_nan=False), encoding="utf-8")
    print(json.dumps({
        "output": str(OUT / "inventory.json"),
        "mat_bytes": sum(item["bytes"] for item in result["mat_sources"]),
        "xlsx_rows": {item["path"]: sum(sheet["rows"] for sheet in item["sheets"]) for item in result["xlsx_exports"]},
        "mat_dataset_counts": {item["path"]: item.get("dataset_count", item["status"]) for item in result["mat_sources"]},
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
