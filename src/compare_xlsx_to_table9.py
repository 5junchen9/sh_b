"""Read-only reconciliation of local XLSX exports to Supplementary Table 9."""
from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "data_audit"
PAPER = OUT / "paper.txt"
ROW_RE = re.compile(
    r"^\s*(EL\d+)\s+(.+?)\s+(20\d\d-\d\d-\d\d)\s+(\d+)\s+(.+?)\s*$",
    re.IGNORECASE,
)


def clean(value):
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def read_sheet_rows(path: Path, wanted: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = next(sheet for sheet in wb.worksheets if sheet.title == wanted)
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(iterator)]
    rows = [dict(zip(headers, (clean(value) for value in values))) for values in iterator if any(value is not None for value in values)]
    wb.close()
    return rows


def official_rows() -> list[dict]:
    records = []
    for line in PAPER.read_text(encoding="utf-8", errors="replace").splitlines():
        match = ROW_RE.match(line)
        if match:
            barcode, dataset, batch_date, cycle_life, policy = match.groups()
            records.append({
                "barcode": barcode.upper(), "dataset": dataset.strip(), "batch_date": batch_date,
                "cycle_life_table9": int(cycle_life), "policy_table9": policy.strip(),
            })
    unique = {row["barcode"]: row for row in records}
    return list(unique.values())


def policy_core(value: str) -> str:
    return re.sub(r"-newstructure$", "", str(value), flags=re.IGNORECASE).replace(" ", "").lower()


def parse_policy(value: str) -> tuple[float, float, float]:
    match = re.fullmatch(r"\s*([\d.]+)C\(([\d.]+)%\)-([\d.]+)C\s*", value)
    if not match:
        raise ValueError(f"Cannot parse Table 9 policy: {value!r}")
    return tuple(float(item) for item in match.groups())


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    official = official_rows()
    if len(official) != 124:
        raise RuntimeError(f"Expected 124 unique Table 9 cells, parsed {len(official)}. Check PDF extraction.")
    roster = {row["barcode"]: row for row in official}
    metas, summaries, vdlin = [], [], []
    for index in range(1, 4):
        book = ROOT / f"data_{index}.xlsx"
        for row in read_sheet_rows(book, "批次元数据"):
            row["source_file"] = book.name
            row["barcode_norm"] = str(row.get("barcode", "")).upper()
            metas.append(row)
        for row in read_sheet_rows(book, "循环汇总"):
            row["source_file"] = book.name
            summaries.append(row)
        for row in read_sheet_rows(book, "Vdlin"):
            row["source_file"] = book.name
            vdlin.append(row)

    by_barcode = defaultdict(list)
    for row in metas:
        by_barcode[row["barcode_norm"]].append(row)
    unique_local = set(by_barcode)
    matched = sorted(unique_local & set(roster))
    local_only = sorted(unique_local - set(roster))
    duplicate_barcodes = {key: rows for key, rows in by_barcode.items() if len(rows) > 1}
    reconciled = []
    for barcode in matched:
        official_row = roster[barcode]
        for local in by_barcode[barcode]:
            local_life = local.get("cycle_life")
            reconciled.append({
                "barcode": barcode,
                "source_file": local["source_file"],
                "batch_index": local.get("batch_index"),
                "local_cycle_life": local_life,
                "dataset": official_row["dataset"],
                "batch_date": official_row["batch_date"],
                "table9_cycle_life": official_row["cycle_life_table9"],
                "cycle_life_match": local_life == official_row["cycle_life_table9"],
                "local_policy": local.get("policy_readable"),
                "table9_policy": official_row["policy_table9"],
                "policy_match_exact": str(local.get("policy_readable")) == official_row["policy_table9"],
                "policy_match_after_newstructure_suffix": policy_core(local.get("policy_readable")) == policy_core(official_row["policy_table9"]),
            })
    life_mismatches = [row for row in reconciled if not row["cycle_life_match"]]
    exact_policy_mismatches = [row for row in reconciled if not row["policy_match_exact"]]
    summary_keys = Counter((row["source_file"], str(row.get("batch_index")), str(row.get("cycle_index"))) for row in summaries)
    summary_duplicate_keys = sum(count - 1 for count in summary_keys.values() if count > 1)
    summary_missing = {field: sum(row.get(field) is None for row in summaries) for field in ["QDischarge", "QCharge", "IR", "Tmax", "Tavg", "Tmin", "chargetime"]}
    summary_counts = Counter((row["source_file"], row.get("batch_index")) for row in summaries)
    vdlin_counts = Counter((row["source_file"], row.get("batch_index")) for row in vdlin)
    record_shape_issues = []
    for row in metas:
        key = (row["source_file"], row.get("batch_index"))
        if summary_counts[key] != row.get("summary_row_count"):
            record_shape_issues.append({"type": "summary_length", "source_file": key[0], "batch_index": key[1], "expected": row.get("summary_row_count"), "actual": summary_counts[key]})
        if vdlin_counts[key] != row.get("vdlin_length"):
            record_shape_issues.append({"type": "vdlin_length", "source_file": key[0], "batch_index": key[1], "expected": row.get("vdlin_length"), "actual": vdlin_counts[key]})
    report = {
        "official_table9_rows": len(official),
        "local_records": len(metas),
        "local_unique_barcodes": len(unique_local),
        "matched_official_barcodes": len(matched),
        "official_barcodes_missing_locally": sorted(set(roster) - unique_local),
        "local_barcodes_not_in_table9": local_only,
        "duplicate_local_barcodes": {key: [{"source_file": row["source_file"], "batch_index": row.get("batch_index")} for row in rows] for key, rows in duplicate_barcodes.items()},
        "cycle_life_mismatch_count": len(life_mismatches),
        "policy_exact_mismatch_count": len(exact_policy_mismatches),
        "cycle_summary_rows": len(summaries),
        "cycle_summary_duplicate_key_count": summary_duplicate_keys,
        "cycle_summary_missing_counts": summary_missing,
        "record_shape_issue_count": len(record_shape_issues),
        "record_shape_issues": record_shape_issues,
        "vdlin_rows": len(vdlin),
        "reconciled_rows": reconciled,
    }
    (OUT / "table9_reconciliation.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    with (OUT / "table9_reconciliation.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(reconciled[0]))
        writer.writeheader()
        writer.writerows(reconciled)
    clean_roster = []
    for barcode in sorted(roster):
        official_row = roster[barcode]
        c1, q1, c2 = parse_policy(official_row["policy_table9"])
        local_rows = by_barcode[barcode]
        clean_roster.append({
            "barcode": barcode,
            "dataset_table9": official_row["dataset"],
            "batch_date_table9": official_row["batch_date"],
            "cycle_life_table9": official_row["cycle_life_table9"],
            "policy_table9": official_row["policy_table9"],
            "C1": c1,
            "Q1_percent": q1,
            "C2": c2,
            "local_fragment_count": len(local_rows),
            "local_source_files": ";".join(row["source_file"] for row in local_rows),
            "local_batch_indices": ";".join(str(row["batch_index"]) for row in local_rows),
        })
    with (OUT / "clean_cell_roster.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(clean_roster[0]))
        writer.writeheader()
        writer.writerows(clean_roster)
    print(json.dumps({key: report[key] for key in report if key not in {"reconciled_rows", "duplicate_local_barcodes", "record_shape_issues"}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
