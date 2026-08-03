"""Create traceable, non-destructive cycle-level modeling tables for L1.

Inputs are the audit-validated XLSX exports and frozen Table 9 roster.  This
script does not alter source files; questionable values are flagged, never
imputed or silently removed.
"""
from __future__ import annotations

import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
AUDIT = ROOT / "outputs" / "data_audit"
OUT = ROOT / "data" / "processed"
CORE = ("QDischarge", "QCharge", "IR", "Tmax", "Tavg", "Tmin", "chargetime")
NONNEGATIVE_FIELDS = ("QDischarge", "QCharge", "IR", "chargetime")


def rows_from_sheet(path: Path, name: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = next(sheet for sheet in workbook.worksheets if sheet.title == name)
    iterator = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(iterator)]
    rows = [dict(zip(header, values)) for values in iterator if any(value is not None for value in values)]
    workbook.close()
    return rows


def basic_field_valid(field: str, value) -> bool:
    if not isinstance(value, (int, float)) or not math.isfinite(value):
        return False
    return field not in NONNEGATIVE_FIELDS or value >= 0


def linear_slope(points: list[tuple[float, float]]) -> float | None:
    if len(points) < 2:
        return None
    xbar = sum(x for x, _ in points) / len(points)
    ybar = sum(y for _, y in points) / len(points)
    denominator = sum((x - xbar) ** 2 for x, _ in points)
    return None if denominator == 0 else sum((x - xbar) * (y - ybar) for x, y in points) / denominator


def flag_is_true(value: str) -> bool:
    """Accept MATLAB CSV logical representations without treating other text as true."""
    return str(value).strip().lower() in {"true", "1"}


def load_and_validate_deep_audit() -> dict | None:
    """Require the cycle mask and its summary to agree before curve work is declared ready."""
    summary_path = AUDIT / "mat_deep_cycle_summary.json"
    flags_path = AUDIT / "mat_deep_cycle_flags.csv"
    if not summary_path.exists() and not flags_path.exists():
        return None
    if not summary_path.exists() or not flags_path.exists():
        raise RuntimeError("MAT deep-audit summary and cycle-mask CSV must either both exist or both be absent.")

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    with flags_path.open(encoding="utf-8-sig", newline="") as stream:
        flags = list(csv.DictReader(stream))
    required = {
        "source_index", "batch_index", "cycle_index", "in_official_roster",
        "usable_for_curve_features", "failure_reason",
    }
    if not flags or not required.issubset(flags[0]):
        raise RuntimeError("MAT deep-audit cycle-mask CSV is missing required columns.")
    keys = [(row["source_index"], row["batch_index"], row["cycle_index"]) for row in flags]
    if len(keys) != len(set(keys)):
        raise RuntimeError("MAT deep-audit cycle-mask CSV has duplicate source/batch/cycle keys.")

    official = [row for row in flags if flag_is_true(row["in_official_roster"])]
    checks = {
        "total_cycles": len(flags),
        "official_roster_cycles": len(official),
        "usable_for_curve_features_count": sum(flag_is_true(row["usable_for_curve_features"]) for row in official),
        "unusable_official_cycle_count": sum(not flag_is_true(row["usable_for_curve_features"]) for row in official),
    }
    mismatches = {name: (summary.get(name), actual) for name, actual in checks.items() if summary.get(name) != actual}
    if mismatches:
        raise RuntimeError(f"MAT deep-audit summary disagrees with cycle-mask CSV: {mismatches}")
    return summary


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    reconciled = json.loads((AUDIT / "table9_reconciliation.json").read_text(encoding="utf-8"))["reconciled_rows"]
    mat_metadata = json.loads((AUDIT / "mat_metadata.json").read_text(encoding="utf-8"))
    deep_audit = load_and_validate_deep_audit()
    source_dates = {source["xlsx_file"]: source["batch_date"] for source in mat_metadata["sources"]}
    fragment_lookup = {(row["source_file"], int(row["batch_index"])): row for row in reconciled}
    labels = list(csv.DictReader((AUDIT / "clean_cell_roster.csv").open(encoding="utf-8-sig")))
    duplicate_barcodes = {row["barcode"] for row in labels if int(row["local_fragment_count"]) > 1}

    cycle_rows = []
    fragment_counts = Counter()
    for source_index in range(1, 4):
        source_name = f"data_{source_index}.xlsx"
        for raw in rows_from_sheet(ROOT / source_name, "循环汇总"):
            key = (source_name, int(raw["batch_index"]))
            label = fragment_lookup.get(key)
            if label is None:
                continue  # 11 local-only barcodes are retained in source, excluded from the frozen roster.
            row = {
                "barcode": label["barcode"],
                "source_file": source_name,
                "batch_index": int(raw["batch_index"]),
                "cycle_index": int(raw["cycle_index"]),
                "dataset_table9": label["dataset"],
                "cycle_life_table9": int(label["table9_cycle_life"]),
                "policy_table9": label["table9_policy"],
                "is_duplicate_barcode": label["barcode"] in duplicate_barcodes,
            }
            row.update({field: raw[field] for field in CORE})
            failed = [field for field in CORE if not basic_field_valid(field, raw[field])]
            row["all_zero_placeholder"] = all(raw[field] == 0 for field in CORE)
            row["core_numeric_valid"] = not failed and not row["all_zero_placeholder"]
            row["core_invalid_fields"] = ";".join(failed)
            cycle_rows.append(row)
            fragment_counts[key] += 1

    # Per-fragment chronological check.  No sorting is performed: sorting could
    # break correspondence with raw cycle arrays.
    grouped = defaultdict(list)
    for row in cycle_rows:
        grouped[(row["source_file"], row["batch_index"])].append(row)
    chronology_issues = {}
    for key, items in grouped.items():
        inversions = sum(items[i]["cycle_index"] <= items[i - 1]["cycle_index"] for i in range(1, len(items)))
        chronology_issues[key] = inversions
        for item in items:
            item["cycle_index_order_ok"] = inversions == 0

    # The five duplicated barcodes are continuation fragments.  Sort by the
    # batch dates read from the source MAT metadata, never by file-name lore.
    cell_groups = defaultdict(list)
    for key, items in grouped.items():
        cell_groups[items[0]["barcode"]].append((key, items))
    merge_issues = []
    merge_valid = {}
    for barcode, fragments in cell_groups.items():
        fragments.sort(key=lambda pair: (source_dates[pair[0][0]], pair[0][1]))
        offset = 0
        expected = fragments[0][1][0]["cycle_life_table9"] - 1
        for order, (_, items) in enumerate(fragments, start=1):
            for item in items:
                item["fragment_order"] = order
                item["global_cycle_index"] = offset + item["cycle_index"]
            offset += len(items)
        if len(fragments) > 1 and offset != expected:
            merge_issues.append({"barcode": barcode, "cycle_rows": offset, "expected_cycle_rows": expected})
        merge_valid[barcode] = len(fragments) == 1 or offset == expected

    # A deliberately conservative, label-free far-outlier flag for charge
    # time.  The threshold is five times the global P99 after excluding
    # structural zero placeholders.  Values are retained in the cycle table,
    # but excluded from preliminary feature aggregation.
    valid_charge_times = sorted(
        float(row["chargetime"])
        for row in cycle_rows
        if row["core_numeric_valid"] and float(row["chargetime"]) > 0
    )
    p99_index = max(0, math.ceil(0.99 * len(valid_charge_times)) - 1)
    chargetime_far_outlier_threshold = 5.0 * valid_charge_times[p99_index]
    for row in cycle_rows:
        row["chargetime_far_outlier"] = (
            row["core_numeric_valid"]
            and float(row["chargetime"]) > chargetime_far_outlier_threshold
        )
        row["feature_eligible"] = row["core_numeric_valid"] and not row["chargetime_far_outlier"]

    fieldnames = list(cycle_rows[0])
    with (OUT / "cycle_summary_clean.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(cycle_rows)
    with (OUT / "cell_labels.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(labels[0]))
        writer.writeheader()
        writer.writerows(labels)

    feature_rows = []
    for barcode, fragments in cell_groups.items():
        items = [item for _, fragment in fragments for item in fragment]
        items.sort(key=lambda item: item["global_cycle_index"])
        early = [item for item in items if item["global_cycle_index"] <= 100 and item["feature_eligible"]]
        first = items[0]
        row = {"barcode": first["barcode"]}
        row.update({"early_cycle_limit": 100, "early_valid_cycle_count": len(early), "fragment_count": len(fragments), "fragment_merge_valid": merge_valid[barcode]})
        for field in CORE:
            values = [float(item[field]) for item in early]
            row[f"{field}_early_mean"] = sum(values) / len(values) if values else None
            row[f"{field}_early_slope_per_cycle"] = linear_slope([(item["global_cycle_index"], float(item[field])) for item in early])
        feature_rows.append(row)
    with (OUT / "early_cycle_features.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(feature_rows[0]))
        writer.writeheader()
        writer.writerows(feature_rows)
    feature_columns = [name for name in feature_rows[0] if name not in {"barcode", "early_cycle_limit", "early_valid_cycle_count", "fragment_count", "fragment_merge_valid"}]
    (OUT / "feature_columns.json").write_text(json.dumps({"id_column": "barcode", "feature_columns": feature_columns, "target_table": "cell_labels.csv", "target_column": "cycle_life_table9", "split_column": "dataset_table9"}, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "overall_readiness": "ready_with_masked_raw_curves" if deep_audit else "summary_baseline_ready_raw_curves_blocked",
        "official_cells": len(labels),
        "local_fragments_included": len(grouped),
        "cell_level_feature_rows": len(feature_rows),
        "cycle_rows_included": len(cycle_rows),
        "all_zero_placeholder_count": sum(row["all_zero_placeholder"] for row in cycle_rows),
        "core_invalid_row_count": sum(not row["core_numeric_valid"] for row in cycle_rows),
        "chargetime_far_outlier_threshold": chargetime_far_outlier_threshold,
        "chargetime_far_outlier_count": sum(row["chargetime_far_outlier"] for row in cycle_rows),
        "raw_curve_audit": deep_audit,
        "fragment_cycle_index_inversion_count": sum(value for value in chronology_issues.values()),
        "fragments_with_cycle_index_inversions": sum(value > 0 for value in chronology_issues.values()),
        "duplicate_merge_issue_count": len(merge_issues),
        "cleaning_policy": "No source values were changed, imputed, or deleted. Structural zero placeholders and conservative charge-time far outliers are retained with flags and excluded only from preliminary feature aggregation.",
        "remaining_risks": [
            "Units for IR and chargetime are not finalized by the supplied metadata.",
            "Five barcodes have multiple source fragments and must remain grouped in all split schemes.",
            "Raw curve features must apply mat_deep_cycle_flags.csv at field-cycle level; 258 official cycles are masked.",
        ],
        "outputs": ["cell_labels.csv", "cycle_summary_clean.csv", "early_cycle_features.csv", "feature_columns.json"],
    }
    (OUT / "data_preparation_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
